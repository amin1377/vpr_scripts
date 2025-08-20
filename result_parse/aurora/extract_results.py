import os
import re
import argparse
from collections import defaultdict
from multiprocessing import Pool, cpu_count
from functools import partial

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError as e:
    raise SystemExit("This script requires 'openpyxl'. Install it with: pip install openpyxl") from e


def parse_config_file(config_filename):
    """Reads the config file and returns a dict: {output_file: [(metric_name, regex_pattern)]}."""
    config_entries = defaultdict(list)
    with open(config_filename, "r") as config_file:
        for line in config_file:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split(";")
            if len(parts) == 3:
                metric_name, output_file, regex_pattern = parts
                print(f"Metric Name: {metric_name} - Regex pattern: {regex_pattern}")
                config_entries[output_file.strip()].append((metric_name.strip(), regex_pattern.strip()))
    return config_entries


def extract_metrics(config_entries, circuit_dir, metrics_map):
    """Extracts metrics from files in circuit_dir into metrics_map."""
    for _, metric_patterns in config_entries.items():
        for metric_name, _ in metric_patterns:
            metrics_map.setdefault(metric_name, None)

    for output_file, metric_patterns in config_entries.items():
        path = os.path.join(circuit_dir, output_file)
        try:
            with open(path, "r") as f:
                for line in f:
                    s = line.strip()
                    for metric_name, regex_pattern in metric_patterns:
                        if metrics_map.get(metric_name) not in (None, -1):
                            continue
                        m = re.search(regex_pattern, s)
                        if m:
                            metrics_map[metric_name] = m.group(1).strip()
        except FileNotFoundError:
            print(f"Warning: File '{path}' not found. Marking related metrics as None.")
            for metric_name, _ in metric_patterns:
                metrics_map[metric_name] = None


def process_circuit(args):
    """
    Process a single circuit. This function will be called by multiprocessing workers.
    Args: tuple of (circuit_name, circuit_dir, config_entries)
    Returns: dict with circuit metrics
    """
    circuit_name, circuit_dir, config_entries = args
    
    print(f"     Processing circuit: {circuit_name} (dir: {circuit_dir})")
    metrics_map = {"circuit": circuit_name}
    extract_metrics(config_entries, circuit_dir, metrics_map)
    return metrics_map


def process_seed_parallel(seed_path, seed_name, config_entries, num_processes):
    """
    Process all circuits in a seed directory using multiprocessing.
    Returns list of circuit metric dictionaries.
    """
    print(f"  Processing seed: {seed_name}")
    
    # Collect circuit processing tasks
    circuit_tasks = []
    for circuit_item in sorted(os.listdir(seed_path)):
        circuit_outer = os.path.join(seed_path, circuit_item)
        if os.path.isdir(circuit_outer):
            circuit_name = circuit_item
            circuit_dir = os.path.join(circuit_outer, circuit_name)
            
            if os.path.isdir(circuit_dir):
                circuit_tasks.append((circuit_name, circuit_dir, config_entries))
            else:
                print(f"    Warning: Expected circuit directory not found: {circuit_dir}")
    
    # Process circuits in parallel
    if circuit_tasks:
        if num_processes > 1 and len(circuit_tasks) > 1:
            # Use multiprocessing for multiple circuits
            with Pool(processes=min(num_processes, len(circuit_tasks))) as pool:
                seed_data = pool.map(process_circuit, circuit_tasks)
        else:
            # Single process fallback
            seed_data = [process_circuit(task) for task in circuit_tasks]
    else:
        seed_data = []
    
    return seed_data


def get_global_metric_keys(all_task_data):
    """Union of metric keys across all tasks and seeds (excluding 'circuit'), sorted."""
    keys = set()
    for task_data in all_task_data:
        for seed_data in task_data.values():
            for row in seed_data:
                keys.update([k for k in row.keys() if k != "circuit"])
    return sorted(keys)


def write_seed_sheet(wb, sheet_name, rows, metric_keys):
    """Write one seed sheet with headers: ['circuit'] + metric_keys."""
    ws = wb.create_sheet(title=sheet_name)
    headers = ["circuit"] + metric_keys
    ws.append(headers)
    for r in rows:
        ws.append([r.get("circuit", "")] + [r.get(k, "") for k in metric_keys])


def get_metric_col_map(wb, sheet_name, metric_keys):
    """Reads a sheet's header row and returns a map of metric names to column letters."""
    if sheet_name not in wb.sheetnames:
        return {}
    
    ws = wb[sheet_name]
    header_row = [cell.value for cell in ws[1]]
    
    metric_col_map = {}
    for col_idx, header in enumerate(header_row, 1):
        if header in metric_keys:
            metric_col_map[header] = get_column_letter(col_idx)
    return metric_col_map


def generate_average_sheets(wb, all_task_data, metric_keys):
    """
    Creates a separate sheet for each task, containing average values for each metric
    across all seeds for that task, using Excel formulas.
    """
    print("Generating average sheets...")
    all_circuits = sorted(set(row.get("circuit") for task_data in all_task_data for seed_data in task_data.values() for row in seed_data))

    # Dynamically get the column map from a sample seed sheet
    first_task_data = all_task_data[0]
    first_seed_name = sorted(first_task_data.keys())[0]
    sample_sheet_name = f"Task_1_{first_seed_name}"
    
    # We must assume the seed sheets have already been written to the workbook
    raw_col_map = get_metric_col_map(wb, sample_sheet_name, metric_keys)

    for task_idx, task_data in enumerate(all_task_data, start=1):
        avg_sheet_name = f"Task_{task_idx}_Avg"
        ws_avg = wb.create_sheet(title=avg_sheet_name)
        headers = ["circuit"] + metric_keys
        ws_avg.append(headers)
        
        for r_idx, circuit in enumerate(all_circuits, start=2):
            ws_avg.cell(row=r_idx, column=1, value=circuit)
            
            for m_idx, metric in enumerate(metric_keys):
                metric_col_letter = raw_col_map.get(metric)
                if not metric_col_letter:
                    continue
                
                # Build list of seed sheet references for this task
                seed_refs = []
                for seed_name in sorted(task_data.keys()):
                    sheet_name = f"Task_{task_idx}_{seed_name}"

                    # FIX: Add ISBLANK check to prevent VALUE() from converting empty cells to 0
                    ref = f'IF(ISBLANK(INDEX(\'{sheet_name}\'!${metric_col_letter}:${metric_col_letter}, MATCH($A{r_idx}, \'{sheet_name}\'!$A:$A, 0))), "", VALUE(INDEX(\'{sheet_name}\'!${metric_col_letter}:${metric_col_letter}, MATCH($A{r_idx}, \'{sheet_name}\'!$A:$A, 0))))'
                    seed_refs.append(ref)
                
                # Check if there are any references to average.
                if seed_refs:
                    # The criterion is ">-1" to include all non-negative numbers
                    avg_formula = f'=IFERROR(AVERAGEIF({{{", ".join(seed_refs)}}}, ">-1"), "")'
                    ws_avg.cell(row=r_idx, column=m_idx + 2, value=avg_formula)
                else:
                    # If no seeds or data found, leave the cell blank
                    ws_avg.cell(row=r_idx, column=m_idx + 2, value="")


def build_ratio_sheet(wb, all_task_data, metric_keys):
    """
    Creates a 'Comparison' sheet with only ratio columns, referencing the average sheets.
    """
    if not all_task_data or len(all_task_data) < 2:
        return 

    ws = wb.create_sheet(title="Comparison")

    # Get the column map for the newly created average sheets
    avg_col_map = get_metric_col_map(wb, "Task_1_Avg", metric_keys)
    if not avg_col_map:
        print("Warning: Task_1_Avg sheet not found. Cannot build comparison sheet.")
        return

    # Collect all circuits across all tasks and seeds
    all_circuits = sorted(set(row.get("circuit") for task_data in all_task_data for seed_data in task_data.values() for row in seed_data))
    num_tasks = len(all_task_data)
    
    # Create headers for ratios only
    headers = ["circuit"]
    for metric in metric_keys:
        for task_idx in range(2, num_tasks + 1):
            headers.append(f"{metric}_ratio_Task{task_idx}_vs_Task1")
    ws.append(headers)

    # Write data rows
    for r_idx, circuit in enumerate(all_circuits, start=2):
        ws.cell(row=r_idx, column=1, value=circuit)  # circuit name
        
        col_ptr = 2
        for metric in metric_keys:
            base_col_letter = avg_col_map.get(metric) # Task 1's average column letter
            if not base_col_letter: continue

            # Reference Task_1_Avg sheet
            base_ref = f'\'Task_1_Avg\'!{base_col_letter}{r_idx}'
            
            for task_idx in range(1, num_tasks): # Task_2, Task_3, etc.
                compare_col_letter = avg_col_map.get(metric)
                if not compare_col_letter: continue

                # Reference Task_i_Avg sheet
                compare_ref = f'\'Task_{task_idx + 1}_Avg\'!{compare_col_letter}{r_idx}'
                
                # Corrected formula with ISBLANK check
                ratio_formula = f'=IF(OR({base_ref}="", {compare_ref}=""), "", {compare_ref}/{base_ref})'

                ws.cell(row=r_idx, column=col_ptr, value=ratio_formula)
                col_ptr += 1


def getArgs():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--task_dir",
        nargs='+',
        required=True,
        help="List of directories containing seed directories with circuits"
    )
    parser.add_argument("--out_file_name", required=True, help="Output .xlsx file name")
    parser.add_argument("--config_file", required=True, help="Config file with metrics/regex")
    parser.add_argument(
        "--num_processes", 
        type=int, 
        default=cpu_count(),
        help=f"Number of parallel processes to use (default: {cpu_count()}, your CPU core count)"
    )
    return parser.parse_args()


def main():
    args = getArgs()
    out_xlsx = args.out_file_name if args.out_file_name.lower().endswith(".xlsx") else f"{args.out_file_name}.xlsx"
    num_processes = args.num_processes

    print(f"Using {num_processes} processes for parallel processing")
    config_entries = parse_config_file(args.config_file)

    # Build all_task_data: list per task_dir; each task is dict per seed; each seed is list of dict rows
    all_task_data = []
    
    for task_idx, task_dir in enumerate(args.task_dir, start=1):
        print(f"Processing task_dir {task_idx}: {task_dir}")
        task_data = {}  # seed_name -> list of circuit rows
        
        # Look for seed directories
        seed_tasks = []
        for item in sorted(os.listdir(task_dir)):
            seed_path = os.path.join(task_dir, item)
            if os.path.isdir(seed_path) and item.startswith("seed_"):
                seed_tasks.append((seed_path, item))
        
        # Process seeds (each seed processes its circuits in parallel)
        for seed_path, seed_name in seed_tasks:
            seed_data = process_seed_parallel(seed_path, seed_name, config_entries, num_processes)
            task_data[seed_name] = seed_data
        
        all_task_data.append(task_data)

    wb = Workbook()
    wb.remove(wb.active)

    # Get consistent metric columns across all tasks and seeds
    metric_keys = get_global_metric_keys(all_task_data)

    # Write individual seed sheets for each task
    for task_idx, task_data in enumerate(all_task_data, start=1):
        for seed_name, seed_data in task_data.items():
            sheet_name = f"Task_{task_idx}_{seed_name}"
            write_seed_sheet(wb, sheet_name, seed_data, metric_keys)

    # Step 1: Generate average sheets for each task
    generate_average_sheets(wb, all_task_data, metric_keys)
    
    # Step 2: Generate comparison sheet with ratios only
    build_ratio_sheet(wb, all_task_data, metric_keys)

    wb.save(out_xlsx)
    print(f"✅ Wrote Excel report with seed-based analysis: {out_xlsx}")


if __name__ == "__main__":
    main()